import requests_cache
from bs4 import BeautifulSoup
from pathlib import Path

import csv
import datetime as dt
from pprint import pprint
import pandas as pd

from openpyxl import Workbook

BASE_DIR = Path(__file__).parent
DATETIME_FORMAT = '%Y-%m-%d_%H-%M-%S'

MAIN_DOC_URL = 'https://audac.eu/eu/products/d/ateo4---wall-speaker-with-clevermount-4inch#section-specifications'

cnt = 0


if __name__ == '__main__':
    # Загрузка веб-страницы с кешированием.
    session = requests_cache.CachedSession()
    response = session.get(MAIN_DOC_URL)
    response.encoding = 'utf-8'
    results = []
    # results_1 = []
    # results_2 = []
    # results_3 = []

    # Создание "супа".
    soup = BeautifulSoup(response.text, features='lxml')

    # таблицу взяли
    all_spec_table = soup.find_all('table', attrs={'class': 'table table-hover mb-6 specification-table'})
    # print(all_spec_table)
    # tr_tags = system_spec_table.find_all('tr')


    # взял вторую таблицу из трех на странице. остальные потом/
    system_spec_table = all_spec_table[1]

    # product_features_table = all_spec_table[2]
    # print(system_spec_table)
    # print(product_features_table)


    soup_2 = BeautifulSoup(system_spec_table.text, features='lxml')

    tr_tags_spec = system_spec_table.find_all('tr')

    # # Проходим tr тэгам
    # for tr_tag_spec in tr_tags_spec:
    #     # print('!!!!!!!!!!!!!!!!!!!!!!!!!!')  # просто разделил чтоб подумать)
    #     print(tr_tag_spec)

    # print(tr_tags_spec)


    # table = []
    # for tr_tag in tr_tags_spec:
    #     row = []
    #     for td_tag in tr_tag.find_all('td'):
    #         colspan = td_tag.attrs.get('colspan', 1)
    #         row.append(td_tag.text or None)
    #         row += [None] * (int(colspan) - 1)
    #     table.append(row)

    # pprint(table)

    table = []
    for tr_tag in tr_tags_spec:
        row = []
        for td_tag in tr_tag.find_all('td'):
            colspan = td_tag.attrs.get('colspan')


            # if colspan == '2':
            #     row += [None] * (int(colspan) - 1)
            #     # print('типо - 2')
            # elif colspan == '1':
            #     row += [None] * (int(colspan) - 1)


            row.append(td_tag.text or None)
            # row.append(td_tag.text)


        table.append(row)


    upd_table = [['Характеристика', '|', 'Значение', '||']]


# 1) 2 ой столбец уходит в 1ый + отступ

#     for row in table:
#         if len(row) == 2:
#             row.insert(1, '|')
#             row.insert(3, '||')

#         elif len(row) > 2:
#             if row[0] != None:
#                 row[0] = (row[0] + ' ' * 10) + row[1]

#             elif row[0] == None:
#                 row.insert(0, row[1])
#                 row.insert(2, row[3])
#                 del row[3:5]
#                 pprint(row)

#             row[1] = '|'
#             row.insert(3, '||')

#         upd_table.append(row)


# 2) 2 ой столбец уходит в 3ий + отступ и становится 2ым

    # for row in table:

    #     if len(row) == 2:
    #         row.insert(1, '|')
    #         row.insert(3, '||')

    #     elif len(row) > 2:
    #         row[1] = (row[1] + ' ' * 10) + row[2]
    #         row.insert(1, '|')
    #         row.insert(3, '||')
    #         del row[4]

    #     upd_table.append(row)
        # pprint(upd_table)


# 3) 2 ой столбец уходит в 1ый + отступ

    for row in table:
        if len(row) == 2:
            row.insert(1, '|')
            row.insert(3, '||')

        elif len(row) > 2:
            if row[0] != None:
                c1 = row[1]
                c2 = row[2]
                row = [row[0]]
                upd_row = [c1, c2]
                upd_table.append(upd_row)


                # row[0] = (row[0] + ' ' * 10) + row[1]

            elif row[0] == None:
                row.insert(0, row[1])
                row.insert(2, row[3])
                del row[3:5]
                pprint(row)

        # pprint(row)
            # row[1] = '|'
            # row.insert(3, '||')

        upd_table.append(row)
        # upd_table.append(upd_row)


    # рабочий
    results_dir = BASE_DIR / 'results'
    results_dir.mkdir(exist_ok=True)

    now = dt.datetime.now()
    now_formatted = now.strftime(DATETIME_FORMAT)
    # file_name = f'{now_formatted}.csv'
    file_name = f'{now_formatted}.xlsx'
    file_path = results_dir / file_name
    df = pd.DataFrame(upd_table)
    df.to_excel('file_path.xlsx', index=0, header=False)





    # wb = Workbook()
    # df = pd.DataFrame(table)
    # ws = wb.create_sheet("Mysheet")
    # # df.to_excel(ws, index=False)



    # model_desc = {}
    # curr_attr = None
    # curr_subattr = None
    # for row in table:
    #     curr_attr = row[0] or curr_attr
    #     curr_subattr = row[1]
    #     value = row[2]
    #     if curr_subattr is None:
    #         if curr_attr not in model_desc:
    #             model_desc[curr_attr] = []
    #         model_desc[curr_attr].append(value)
    #     else:
    #         if curr_attr not in model_desc:
    #             model_desc[curr_attr] = {}
    #         model_desc[curr_attr][curr_subattr] = value


    # pprint(model_desc)



    # results_dir = BASE_DIR / 'results'
    # results_dir.mkdir(exist_ok=True)

    # now = dt.datetime.now()
    # now_formatted = now.strftime(DATETIME_FORMAT)
    # # file_name = f'{now_formatted}.csv'
    # file_name = f'{now_formatted}.xlsx'
    # file_path = results_dir / file_name
    # df = pd.DataFrame(table)
    # df.to_excel('file_path.xlsx', index=False)

    # with open(file_path, 'w', encoding='utf-8') as f:
    #     writer = csv.writer(f, delimiter=',', dialect='unix')
    #     # for i in results:
    #     writer.writerows(table)


    # рабочий
    # results_dir = BASE_DIR / 'results'
    # results_dir.mkdir(exist_ok=True)

    # now = dt.datetime.now()
    # now_formatted = now.strftime(DATETIME_FORMAT)
    # # file_name = f'{now_formatted}.csv'
    # file_name = f'{now_formatted}.xlsx'
    # file_path = results_dir / file_name
    # df = pd.DataFrame(upd_table)
    # df.to_excel('file_path.xlsx', index=0, header=False)






# def create_sheet(table):
#     if not wb.active:
#         ws = wb.create_sheet("Mysheet")
#         df = pd.DataFrame(table)
#         df.to_excel(ws, index=False)












    # soup_2 = BeautifulSoup(system_spec_table.text, features='lxml')

    # tr_tags_spec = system_spec_table.find_all('tr')
    # for tr_tag_spec in tr_tags_spec:

    #     # tr_tag_column =
    #     # print(tr_tag_spec.text)
    #     # results.append((tr_tag.text))

    #     # td_tag_spec = tr_tag_spec.find('td')
    #     # td_tag_colspan = tr_tag_spec.find_all('td colspan="2"')
    #     # td_tag_colspan = td_tag_spec.find_next()


    #     # print(td_tag_spec.text)
    #     # print(td_tag_colspan)

    #     # = dt_tag.find_next_sibling().string

    #     # results_1.append(td_tag_spec.text)
    #     # print(results)
    #     results.append(tr_tag_spec.text)







    # results_dir = BASE_DIR / 'results'
    # results_dir.mkdir(exist_ok=True)

    # now = dt.datetime.now()
    # now_formatted = now.strftime(DATETIME_FORMAT)
    # file_name = f'{now_formatted}.csv'
    # file_path = results_dir / file_name
    # with open(file_path, 'w', encoding='utf-8') as f:
    #     writer = csv.writer(f, delimiter=' ')
    #     # for i in results:
    #     writer.writerow(results)
    #         # writer.writerow([i])


    # print(results)

    # # print(results_1)
    # # print(results_2)
    # # print(results_3)


